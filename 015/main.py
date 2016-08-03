#!/usr/bin/env python
# -*- coding: utf-8 -*-

import json
from collections import OrderedDict
from openpyxl import Workbook

def txt_to_xlsx(filename):
    file = open(filename, 'r')
    file_content = json.load(file, encoding = 'UTF-8')
    print(file_content)
    
    workbook = Workbook()
    worksheet = workbook.worksheets[0]
    
    for i in range(1, len(file_content) + 1):
        worksheet.cell(row = i, column = 1).value = i
        worksheet.cell(row = i, column = 2).value = file_content[str(i)]
        
    workbook.save(filename = 'city.xls')

if __name__ == '__main__':
    txt_to_xlsx('city.txt')