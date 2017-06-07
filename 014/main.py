#!/usr/bin/env python
# -*- coding: utf-8 -*-

import json
from collections import OrderedDict
from openpyxl import Workbook

def txt_to_xlsx(filename):
    file = open(filename, 'r')
    file_content = json.load(file)
    print(file_content)
    
    workbook = Workbook()
    worksheet = workbook.worksheets[0]
    
    for i in range(1, len(file_content) + 1):
        worksheet.cell(row = i, column = 1).value = i
        for m in range(0, len(file_content[str(i)])):
            worksheet.cell(row = i, column = m + 2).value = file_content[str(i)][m]
            print(file_content[str(i)][m])
    
    workbook.save(filename = 'student.xls')

if __name__ == '__main__':
    txt_to_xlsx('student.txt')